import os
import openpyxl


class WriteExcel(object):
    """Writes tabular data from list of lists (row by column) to Excel Spreadsheet"""
    def __init__(self, filename):
        """Args:
            filename - filename of Excel file. If file exists it will be
                opened, otherwise new file will be created.
        """
        #Define cell styles
        self.styles={}
        self.styles["head"] = self.make_style(bold_text=True, bg_colour="FF9EB9E5")
        self.styles["body"] = self.make_style()
        # Spreadhsheet filename
        self.filename = filename
        # Open workbook - either existing or new
        if os.path.exists(filename):
            self.wb = openpyxl.load_workbook(filename=filename)
        else:
            self.wb = openpyxl.Workbook()

    def write_block(self, rows, top_row=1, left_col=1, tab_name="Statements", highlight_rows=()):
        """Write block of data to location in spreadsheet.

        Args:
            rows - list of lists containing data for each row (row by column)
            top_row - topmost row to start writing from (min 1)
            left_col - leftmost column to start writing from (as number, min 1)
            tab_name - name of tab to write to (optional).
                       Will be created if not already present.
                       Defaults to currently active tab.
            highlight_rows - list of row indexes of rows to receive
                             bolder hightlighing, to make top row distinctive
                             use [0]
        """
        # If tab name supplied, use this tab (create if necessary)
        if tab_name:
            if tab_name not in self.wb.sheetnames:
                self.wb.create_sheet(tab_name)
            ws = self.wb[tab_name]
        # Default to the active sheet
        else:
            ws = self.wb.active

        # Write the data
        for ri, row in enumerate(rows):
            # Cell Style
            style = self.styles["body"]
            # Set different style for row if it's to be highlighted
            if ri in highlight_rows:
                style = self.styles["body"]
            # Write row to spreadsheet
            for ci, col in enumerate(row):
                cell = ws.cell(row=top_row+ri, column=left_col+ci)
                cell.value = str(col)
                cell.style = style

    def save(self):
        """Save spreadsheet"""
        self.wb.save(self.filename)

    def make_style(self,
                   text_size=10,
                   text_colour="FF000000",
                   bold_text=False,
                   bg_colour=None,
                   border=True):
        """Make and return an openpyxl style"""
        style = openpyxl.styles.named_styles.NamedStyle()
        # Font
        style.font = openpyxl.styles.Font(bold=bold_text, size=text_size, color=text_colour)
        # Background fill
        if bg_colour:
            style.fill = openpyxl.styles.PatternFill(fill_type="solid",
                                                     start_color=bg_colour,
                                                     end_color=bg_colour)
        # Border
        if border:
            bd = openpyxl.styles.Side(style='thin', color="000000")
        else:
            bd = openpyxl.styles.Side(border_style=None)
        style.border = openpyxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
        return style
