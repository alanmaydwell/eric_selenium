#!/usr/bin/env python

"""
Measures transaction times in Eric using some data from a specially
formatted spreadsheet. Also records results to this spreadsheet.

Relies on eric.py to interact with Eric.
"""

# More secure password input
import getpass
import time

# Command-line argument handling
import sys
import os

# Used to read spreadsheet
import openpyxl

# Generic function timer
from eric import fn_timer
# import eric Webdriver handling
from eric import Eric

#Write results to Excel files
import excel_block_write


class ExcelRun(object):
    def __init__(self, filename=""):
        """
        Uses Selenium to search, select and view Reports in Eric based on
        data from specially formatted spreadsheet.
        Args:
            filename - Excel file with test data
        """
        # Selenium runner in Eric
        self.eric = Eric()
        # Excel filename
        self.filename = filename
        # Row where columns headings are located
        self.heading_row = 5
        # When True, export details fro any viewed report to Excel file(s)
        # Value can be changed by flag read from spreadsheet data (in self.run)
        self.extract_details = False
        # Start date/time (used in results filename creation)
        self.run_start = time.strftime("%Y.%m.%d_%H.%M.%S")

        #Sub folder for results - setup if not present
        self.results_folder = os.path.join(os.getcwd(),"extracted_details")
        #Create the folder if it doesn't exist
        if not os.path.exists(self.results_folder):
            os.makedirs(self.results_folder)

    def run(self, max_scenario_row=60):
        """Run using data from spreadsheet

        Args:
            max_scenario_row - maximum row number for scenario. Execution will
                stop after maximum row reached.
        """
        # Try to open spreadsheet
        self.open_spreadsheet()

        # Check that required tabs are present and give up if any missing
        essential_tabs = ["Scenario", "Results"]
        missing_tabs = [tab for tab in essential_tabs if tab not in self.wb.sheetnames]
        if missing_tabs:
            print "Excel file:", self.filename
            print "Ending - required tab(s) missing: " + ", ".join(missing_tabs)
            return "Missing tabs: "+", ".join(missing_tabs)

        # The scenario sheet
        ss = self.wb.get_sheet_by_name("Scenario")
        # The results sheet
        rs = self.wb.get_sheet_by_name("Results")

        # Get the starting row to write_results from
        # Read value recorded in spreadsheet but ensure
        # it is below the heading row
        results_start_row = rs["C3"].value
        if results_start_row <= self.heading_row:
            results_start_row = self.heading_row + 1

        # update self.extract_details based on value from spreadsheet
        # Controls report data
        self.extract_details = str(ss["D3"].value)[:4].lower()

        # Sheet starting results column
        results_start_column = 2

        # Iterate through the scenario rows
        scenario_row = self.heading_row + 1
        results_row = results_start_row
        results_column = results_start_column
        continue_run = True
        while continue_run:
            # Read scenario step from spreadsheet
            action = ss.cell(row=scenario_row, column=1).value
            # make action lower-case str
            action = str(action).lower()
            # Associated parameters
            parameter = ss.cell(row=scenario_row, column=2).value
            # Parameter 2 and 3 currently only used for username/password
            parameter2 = ss.cell(row=scenario_row, column=3).value
            parameter3 = ss.cell(row=scenario_row, column=4).value

            # Record run time to spreadsheet
            rs.cell(row=results_row, column=1).value = time.strftime("%d/%m/%Y - %H:%M:%S")

            # Take action based on step details

            # Login to Eric
            if action == "login":
                # Get username and password if not in spreadsheet
                if not parameter2:
                    parameter2 = raw_input("Username:")
                if not parameter3:
                    parameter3 = getpass.getpass(prompt="Password:")

                rs.cell(row=results_row, column=results_column).value = parameter + " " + parameter2
                results_column += 1
                launch_time = self.check_login(parameter2, parameter3, parameter)
                rs.cell(row=results_row, column=results_column).value = launch_time
                results_column += 1
                # Give up if login was unsuccessful
                if str(launch_time) == "Login Failed":
                    continue_run = False

            elif action == "dlogin":
                rs.cell(row=results_row, column=results_column).value = parameter + " " + parameter2
                self.eric.login_direct(parameter, parameter2)
                results_column += 2

            # Perform search
            elif action == "search":
                rs.cell(row=results_row, column=results_column).value = parameter
                results_column += 1
                search_time = self.check_search(parameter)
                rs.cell(row=results_row, column=results_column).value = search_time
                results_column += 1

            # Select report
            elif action == "select":
                # Record parameter value
                rs.cell(row=results_row, column=results_column).value = parameter
                results_column += 1
                # Check there are reports available to view
                message, report_names = self.eric.report_list_items()
                # Get selection time if reports are present
                if report_names:
                    select_time = self.check_report_select(parameter)
                    rs.cell(row=results_row, column=results_column).value = select_time
                # Write message if reports absent
                else:
                    rs.cell(row=results_row, column=results_column).value = "No Reports Present"
                results_column += 1

            # View report (previously selected)
            elif action == "view":
                # See which report is currently selected
                report_name = self.eric.read_report_choice()
                # Write report name to spreadsheet
                rs.cell(row=results_row, column=results_column).value = report_name
                results_column += 1
                # If we've a report measure the time it takes to open it
                if report_name:
                    view_time = self.check_report_view()
                else:
                    view_time = "n/a - no report selected"
                rs.cell(row=results_row, column=results_column).value = view_time
                results_column += 1

            # Logout from Eric and end Webdriver session
            elif action == "logout":
                # Logout from Eric
                self.eric.log_out()
                # Close Webdriver Session
                self.eric.close()

            # Start new results line
            elif action == "newline":
                results_row += 1
                results_column = results_start_column

            # Advance to next row for next run
            scenario_row += 1

            # Quit if we're at end
            if scenario_row > max_scenario_row or not action:
                continue_run = False

        # Final actions
        # Update the details of the row reached
        results_row += 1
        rs["C3"].value = results_row

        # Save at end
        self.wb.save(self.filename)

        # Quit webdriver (rely on logout action above instead)
        ##self.eric.close()

    def check_login(self, username, password, url):
        """Login to CCR"""
        login_response = self.eric.login(username, password, url)
        if login_response == 0:
            launch_time = fn_timer(self.eric.open_eric)
        else:
            launch_time = "Login Failed"
        return launch_time

    def check_search(self, account_no):
        """Complete Eric search using supplied account number
        and return the time taken for search to conclude.
        """
        search_time = fn_timer(self.eric.search, account_no)
        return search_time

    def check_report_select(self, report_id):
        """Selects one report, times how long page updates with response
        Args:
            report_id - report number on screen, int from 1 to 4
        Returns:
            response time in seconds
        """
        report_no = int(report_id)-1
        if report_no in (0, 1, 2, 3):
            select_time = fn_timer(self.eric.select_report, report_no)
        else:
            select_time = "Invalid report_id '{}'".format(report_id)
        return select_time

    def check_report_view(self):
        """Response time for report to appear after clicking view button"""
        #Read the current report choice before viewing
        # Also capture "N reports for supplier X" message
        current_choice = self.eric.read_report_choice()
        short_name = current_choice.split(" ")[0]
        supplier, _ = self.eric.report_list_items()
        # View the report
        view_time = fn_timer(self.eric.view_report)

        # Write financial details to Excel file or HTML file if flag set
        # Setup some heading info to write to spreadsheet
        url = self.eric.driver.current_url
        now = time.strftime("%d/%m/%Y - %H:%M:%S")
        info = [[supplier, now, url]]
        # Read details from the report
        if self.extract_details in ("html", "both"):
            # Get report content as page source
            details = self.eric.get_report_details()
            self.html_report_write(info, details["source"], short_name+"_source_")
        if self.extract_details in ("exce", "both"):
            # Get report content as individual cells
            details = self.eric.get_report_details(get_cells=True)
            self.excel_report_write(info, details["cells"], short_name)

        # Close the Eric report after viewing (not timed currently)
        self.eric.close_report()
        return view_time

    def html_report_write(self, info, details, report_name):
        """Write extracted financial statement content to an HTML file
        Args:
            info - list containing info text for first line
            details - page source of report HTML tables
            report_name - short name for Financial statement
                         (added to filename).
        """
        report_filename = report_name + self.run_start + ".html"
        html_path = os.path.join(self.results_folder, report_filename)
        with open(html_path, "w") as html_file:
            # Write info at top (between <pre> tags)
            html_file.write("<pre>\n")
            html_file.write("Eric Report Extracted Using Selenium Script</b>\n")
            for item in info[0]:
                html_file.write(item + "\n")
            # Write the source
            html_file.write("</pre> <hr>\n\n")

            # Write each report contents to the file
            html_file.write(details+"\n")

    def excel_report_write(self, info, details, tab_name="Report"):
        """Write extracted financial statement content to an Excel file
        Args:
            info - list containing information values to record in top row
            details - list of lists containing tabular report content
                    (as extracted by self.eric.examine_report_details)
            tab_name - (optional) name of Excel tab to write to
        """
        excel_filename = "Reports_" + self.run_start + ".xlsx"
        excel_path = os.path.join(self.results_folder, excel_filename)
        excel = excel_block_write.WriteExcel(excel_path)
        # Write info text to top row
        excel.write_block(info, top_row=1, tab_name=tab_name, highlight_rows=[0])
        # Write report details to further rows
        start_row = 3
        for item in details:
            excel.write_block(item, start_row, tab_name=tab_name)
            excel.save()
            start_row = start_row + 1 + len(item)

    def open_spreadsheet(self):
        """Open test data spreadsheet"""
        try:
            self.wb = openpyxl.load_workbook(filename=self.filename)
        # Give up if fails
        except Exception as e:
            response = self.filename+" - Failed to read: " + e.__doc__
            print response
            print os.getcwd()
            # Give up if unsuccessful
            return response


if __name__ == "__main__":

    # Default filename
    filename = "eric_data.xlsx"
    # Get filename from command-line arg if supplied
    if len(sys.argv) > 1:
        filename = sys.argv[1]

    go = ExcelRun(filename=filename)
    go.run()
    print "Finished"
